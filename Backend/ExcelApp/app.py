import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import mysql.connector
from mysql.connector import Error

# Define the column mapping
column_mapping = {
    'Customer Name': 'Stockist_Name',
    'Bill No': 'Bill_No',
    'Bill Date': 'Bill_Date',
    'Cust Code': 'Chemist_Code',
    'Product Name': 'Material_Name',
    'Qty': 'Sale_Qty',
    'Free': 'Free_Qty',
    'Amount': 'Value',
    'Batch No': 'Batch_No',
    'Address': 'Address',
    'Area Name': 'City',
    'Pin Code': 'Pin_Code',
    'Rate': 'Rate',
    'Goods Value': 'Value'
}

# Function to validate the columns against the mapping
def validate_columns(df):
    missing_columns = []
    for excel_col in column_mapping.keys():
        if excel_col not in df.columns:
            missing_columns.append(excel_col)
    return missing_columns

# Function to extract tables from Excel and display data
def extract_excel_tables(file):
    try:
        # Load the Excel file using openpyxl
        workbook = load_workbook(file)
        sheet_names = workbook.sheetnames

        # Select a sheet to extract tables from
        selected_sheet = sheet_names[0]
        
        # Load the selected sheet
        sheet = workbook[selected_sheet]
        
        # Extract the tables (if any) within the selected sheet
        if sheet.tables:
            table_names = list(sheet.tables.keys())
            selected_table = table_names[0]
            
            # Extract the table range
            table_range = sheet.tables[selected_table].ref
            
            # Extract the cell values from the table range
            data = sheet[table_range]
            table_data = [[cell.value for cell in row] for row in data]

            # Create a DataFrame using the extracted data
            df = pd.DataFrame(table_data[1:], columns=table_data[0])  # Use the first row as column headers
            
            # Validate columns
            missing_columns = validate_columns(df)
            if missing_columns:
                st.warning(f"Missing columns in the uploaded table: {', '.join(missing_columns)}")
            else:
                # Drop rows with any NaN values
                df.dropna(inplace=True)
                
                if df.empty:
                    st.warning("No valid rows to insert after removing rows with missing values.")
                    return
                
                # Display table data
                st.write(f"Data from table '{selected_table}':")
                st.dataframe(df)

                # Convert 'Bill Date' to the correct format
                df['Bill Date'] = pd.to_datetime(df['Bill Date'], format='%d/%m/%Y', errors='coerce').dt.strftime('%Y-%m-%d')

                # Insert data into MySQL
                if st.button("Insert Data into MySQL"):
                    # Define MySQL connection parameters
                    db_config = {
                        'host': 'localhost',
                        'user': 'root',
                        'password': '',
                        'database': 'apex'
                    }

                    # Prepare SQL insert query
                    insert_query = """
                    INSERT INTO salesdata (Stockist_Code, Stockist_Name, Bill_No, Bill_Date, Chemist_Code, 
                                                  Chemist_Name, Address, City, Pin_Code, Material_Code, 
                                                  Material_Name, Batch_No, Sale_Qty, Free_Qty, Rate, Value)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """
                    
                    # Track the number of rows successfully inserted
                    rows_inserted = 0
                    
                    try:
                        # Connect to MySQL
                        connection = mysql.connector.connect(**db_config)
                        cursor = connection.cursor()

                        # Loop through the DataFrame rows
                        for index, row in df.iterrows():
                            # Prepare values based on the mapping
                            values = (
                                "1232",  # Adjust based on your actual Excel column names
                                "KAMALAM MEDICAL CORPORATION",
                                row.get('Bill No'),
                                row.get('Bill Date'),  # This will now be in the correct format
                                row.get('Cust Code'),  # Ensure to fetch actual chemist code
                                row.get('Customer Name'),  # Ensure to fetch actual chemist name
                                row.get('Address'),
                                row.get('Area Name'),
                                row.get('Pin Code'),
                                'Material Code',  # Ensure to fetch actual material code
                                row.get('Product Name'),
                                row.get('Batch No'),
                                row.get('Qty'),
                                row.get('Free'),
                                row.get('Rate'),
                                row.get('Amount'),
                            )
                            
                            cursor.execute(insert_query, values)
                            rows_inserted += 1

                        # Commit the transaction after all rows
                        connection.commit()
                        st.success(f"Successfully inserted {rows_inserted} rows into MySQL!")
                    except Error as e:
                        st.error(f"Error occurred: {e}")
                    finally:
                        if connection.is_connected():
                            cursor.close()
                            connection.close()
        else:
            st.write(f"No tables found in the '{selected_sheet}' sheet.")
        
    except Exception as e:
        st.error(f"Error occurred: {e}")

# Main function for Streamlit app
def main():
    st.title("Excel Table Extractor and SQL Inserter")
    
    # File uploader in Streamlit
    file = st.file_uploader("Upload your Excel file", type=["xlsx"])

    # Extract and display tables when a file is uploaded
    if file is not None:
        extract_excel_tables(file)

if __name__ == "__main__":
    main()
