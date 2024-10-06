import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import mysql.connector
from mysql.connector import Error

def validate_columns(df):
    column_mapping = {
        'Customer Name': 'Stockist_Name',
        'Bill No': 'Bill_No',
        'Bill Date': 'Bill_Date',
        'Cust Code': 'Chemist_Code',
        'Product Name': 'Material_Name',
        'Qty': 'Sale_Qty',
        'Free': 'Free_Qty',
        'Amount': 'Value',
        'Batch No': 'Batch_No',
        'Address': 'Address',
        'Area Name': 'City',
        'Pin Code': 'Pin_Code',
        'Rate': 'Rate',
        'Goods Value': 'Value'
    }
    missing_columns = [excel_col for excel_col in column_mapping.keys() if excel_col not in df.columns]
    return missing_columns

def extract_excel_tables(file):
    try:
        workbook = load_workbook(file)
        sheet = workbook[workbook.sheetnames[0]]
        
        if sheet.tables:
            table = sheet.tables[list(sheet.tables.keys())[0]]
            data = sheet[table.ref]
            table_data = [[cell.value for cell in row] for row in data]
            df = pd.DataFrame(table_data[1:], columns=table_data[0])
            
            missing_columns = validate_columns(df)
            if missing_columns:
                st.warning(f"Missing columns in the uploaded table: {', '.join(missing_columns)}")
            else:
                df.dropna(inplace=True)
                if df.empty:
                    st.warning("No valid rows to insert after removing rows with missing values.")
                    return None
                
                df['Bill Date'] = pd.to_datetime(df['Bill Date'], format='%d/%m/%Y', errors='coerce').dt.strftime('%Y-%m-%d')
                return df
        else:
            st.write("No tables found in the sheet.")
            return None
    except Exception as e:
        st.error(f"Error occurred: {e}")
        return None

def insert_excel_data_into_mysql(df):
    db_config = {
        'host': 'localhost',
        'user': 'root',
        'password': '',
        'database': 'apex'
    }
    
    insert_query = """
    INSERT INTO salesdata (Stockist_Code, Stockist_Name, Bill_No, Bill_Date, Chemist_Code, 
                          Chemist_Name, Address, City, Pin_Code, Material_Code, 
                          Material_Name, Batch_No, Sale_Qty, Free_Qty, Rate, Value)
    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """
    
    rows_inserted = 0
    
    try:
        connection = mysql.connector.connect(**db_config)
        cursor = connection.cursor()

        for index, row in df.iterrows():
            values = (
                "1232", "KAMALAM MEDICAL CORPORATION", row.get('Bill No'), row.get('Bill Date'),
                row.get('Cust Code'), row.get('Customer Name'), row.get('Address'),
                row.get('Area Name'), row.get('Pin Code'), 'Material Code', row.get('Product Name'),
                row.get('Batch No'), row.get('Qty'), row.get('Free'), row.get('Rate'), row.get('Amount')
            )
            cursor.execute(insert_query, values)
            rows_inserted += 1

        connection.commit()
        st.success(f"Successfully inserted {rows_inserted} rows into MySQL!")
    except Error as e:
        st.error(f"Error occurred: {e}")
    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()