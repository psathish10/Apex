import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import mysql.connector
from mysql.connector import Error
import pdfplumber
import re

# Define the column mapping for Excel
column_mapping = {
    'Customer Name': 'Stockist_Name',
    'Bill No': 'Bill_No',
    'Bill Date': 'Bill_Date',
    'Cust Code': 'Chemist_Code',
    'Product Name': 'Material_Name',
    'Qty': 'Sale_Qty',
    'Free': 'Free_Qty',
    'Amount': 'Value',
    'Batch No': 'Batch_No',
    'Address': 'Address',
    'Area Name': 'City',
    'Pin Code': 'Pin_Code',
    'Rate': 'Rate',
    'Goods Value': 'Value'
}

# SQL database connection details
DB_CONFIG =  {
                        'host': 'srv1508.hstgr.io',
                        'user': 'u840048117_Apex_demo',
                        'password': 'Toolfe@min10!',
                        'database': 'u840048117_Apex_demo'
                    }

# Define the fixed column names for PDF processing
FIXED_COLUMNS = [
    "S.No", "BillNo/Date", "Pharmacy Name", "City", "Batch", "Expiry", "Qty",
    "Free", "Rpl", "PTS", "PTR", "MRP", "TotSales"
]

# Global variable to hold combined DataFrame for PDF
combined_df = pd.DataFrame()

# Function to validate the columns against the mapping
def validate_columns(df):
    missing_columns = []
    for excel_col in column_mapping.keys():
        if excel_col not in df.columns:
            missing_columns.append(excel_col)
    return missing_columns

# Function to extract tables from Excel and display data
def extract_excel_tables(file):
    try:
        workbook = load_workbook(file)
        sheet_names = workbook.sheetnames
        selected_sheet = sheet_names[0]
        sheet = workbook[selected_sheet]
        
        if sheet.tables:
            table_names = list(sheet.tables.keys())
            selected_table =  table_names[0]
            table_range = sheet.tables[selected_table].ref
            
            data = sheet[table_range]
            table_data = [[cell.value for cell in row] for row in data]
            df = pd.DataFrame(table_data[1:], columns=table_data[0])  # Use first row as headers
            
            missing_columns = validate_columns(df)
            if missing_columns:
                st.warning(f"Missing columns in the uploaded table: {', '.join(missing_columns)}")
            else:
                st.write(f"Data from table '{selected_table}':")
                st.dataframe(df)

                # Convert 'Bill Date' to the correct format
                df['Bill Date'] = pd.to_datetime(df['Bill Date'], format='%d/%m/%Y', errors='coerce').dt.strftime('%Y-%m-%d')

                # Insert data into MySQL
                if st.button("Insert Data into MySQL"):
                    push_data_to_database(df, "excel")

        else:
            st.write(f"No tables found in the '{selected_sheet}' sheet.")
        
    except Exception as e:
        st.error(f"Error occurred: {e}")

def clean_and_fix_columns(data):
    df = pd.DataFrame(data)
    
    if len(df.columns) != len(FIXED_COLUMNS):
        df = df.iloc[:, :len(FIXED_COLUMNS)]
        df.columns = FIXED_COLUMNS[:len(df.columns)]
    else:
        df.columns = FIXED_COLUMNS
    
    df = df[~df.apply(lambda row: row.tolist() == FIXED_COLUMNS, axis=1)]
    df = df.dropna(how='any')
    df = df[df.apply(lambda row: all(row.astype(str).str.strip() != ''), axis=1)]
    df = df[~df.apply(lambda row: "Pharmacy Name" in row.values, axis=1)]
    
    if "BillNo/Date" in df.columns:
        df[['Bill No', 'Bill Date']] = df['BillNo/Date'].str.split('/', n=1, expand=True)
        df['Bill Date'] = df['Bill Date'].str.strip()
        df['Bill No'] = df['Bill No'].str.strip()
    
    df = df.drop(columns=['BillNo/Date'], errors='ignore')

    if "City" in df.columns:
        df[['City', 'Pin Code']] = df['City'].str.extract(r'([A-Za-z\s]+)(\d{4,6})$')
    
    return df

def associate_product_names_by_batch(df, product_names):
    result = []
    current_product = None
    current_batch_prefix = None

    for _, row in df.iterrows():
        batch = str(row['Batch'])
        batch_prefix = batch[:3] if len(batch) >= 3 else batch

        if batch_prefix != current_batch_prefix:
            if product_names:
                current_product = product_names.pop(0)
            else:
                current_product = "Unknown Product"
            current_batch_prefix = batch_prefix

        row['Product Name'] = current_product
        result.append(row)

    return pd.DataFrame(result)

def extract_product_name(page):
    text = page.extract_text()
    match = re.search(r"Product Name\s*:\s*(.*)", text)
    return match.group(1).strip() if match else None

def extract_and_combine_tables(pdf_file):
    global combined_df
    combined_df = pd.DataFrame()  # Reset combined DataFrame
    product_names = []

    try:
        with pdfplumber.open(pdf_file) as pdf:
            for page in pdf.pages:
                product_name = extract_product_name(page)
                if product_name:
                    product_names.append(product_name)
                
                extracted_tables = page.extract_tables()
                for table in extracted_tables:
                    df = clean_and_fix_columns(table)
                    if not df.empty:
                        combined_df = pd.concat([combined_df, df], ignore_index=True)

        if not combined_df.empty:
            combined_df = associate_product_names_by_batch(combined_df, product_names)
            st.write("Combined table from all pages with associated product names (based on Batch):")
            st.dataframe(combined_df)

            # Insert button for PDF data
            if st.button("Insert PDF Data into MySQL"):
                push_data_to_database(combined_df, "pdf")

        else:
            st.write("No valid data found in the PDF tables.")
    
    except Exception as e:
        st.error(f"Error occurred while extracting tables: {str(e)}")

def push_data_to_database(df, source):
    try:
        connection = mysql.connector.connect(**DB_CONFIG)
        cursor = connection.cursor()

        insert_query = """
        INSERT INTO salesdata (Stockist_Code, Stockist_Name, Bill_No, Bill_Date, Chemist_Code,
                                       Chemist_Name, Address, City, Pin_Code, Material_Code, 
                                       Material_Name, Batch_No, Sale_Qty, Free_Qty, Rate, Value)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """

        for index, row in df.iterrows():
            if source == "excel":
                values = (
                    row.get('Cust Code') if pd.notna(row.get('Cust Code')) else None,
                    "KAMALAM MEDICAL CORPORATION",
                    row.get('Bill No') if pd.notna(row.get('Bill No')) else None,
                    row.get('Bill Date') if pd.notna(row.get('Bill Date')) else None,
                    'Chemist Code',
                    row.get('Customer Name') if pd.notna(row.get('Customer Name')) else None,
                    row.get('Address') if pd.notna(row.get('Address')) else None,
                    row.get('Area Name') if pd.notna(row.get('Area Name')) else None,
                    row.get('Pin Code') if pd.notna(row.get('Pin Code')) else None,
                    'Material Code',
                    row.get('Product Name') if pd.notna(row.get('Product Name')) else None,
                    row.get('Batch No') if pd.notna(row.get('Batch No')) else None,
                    row.get('Qty') if pd.notna(row.get('Qty')) else None,
                    row.get('Free') if pd.notna(row.get('Free')) else None,
                    row.get('Rate') if pd.notna(row.get('Rate')) else None,
                    row.get('Amount') if pd.notna(row.get('Amount')) else None,
                )
            else:  # PDF
                bill_date = pd.to_datetime(row['Bill Date'], format='%d-%m-%Y', errors='coerce')
                bill_date_str = bill_date.strftime('%Y-%m-%d') if pd.notna(bill_date) else None

                values = (
                    row['S.No'] if pd.notna(row['S.No']) else None,
                    row['Pharmacy Name'] if pd.notna(row['Pharmacy Name']) else None,
                    row['Bill No'] if pd.notna(row['Bill No']) else None,
                    bill_date_str,
                    "Chemist Code",
                    "Chemist Name",
                    "Address",
                    row['City'] if pd.notna(row['City']) else None,
                    row['Pin Code'] if pd.notna(row['Pin Code']) else None,
                    "Material Code",
                    row['Product Name'] if pd.notna(row['Product Name']) else None,
                    row['Batch'] if pd.notna(row['Batch']) else None,
                    row['Qty'] if pd.notna(row['Qty']) else None,
                    row['Free'] if pd.notna(row['Free']) else None,
                    row['PTR'] if pd.notna(row['PTR']) else None,
                    row['TotSales'] if pd.notna(row['TotSales']) else None
                )

            cursor.execute(insert_query, values)

        connection.commit()
        st.success("Data has been successfully inserted into the database.")

    except mysql.connector.Error as err:
        st.error(f"Error: {err}")
    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()

def main():
    # Display the company logo in the sidebar
   
    
    st.title("Apex Lab's Data Processor - Demo")
    
    # Sidebar for file upload
    st.sidebar.header("Upload File")
    uploaded_file = st.sidebar.file_uploader("Upload an Excel or PDF file", type=["xlsx", "pdf"])

    if st.sidebar.button("See Your Data"):
        # Replace with the actual link to your admin panel
        st.sidebar.markdown("[Go to Admin Panel](https://toolfe.com/apexdemo/dashboard/index.php)", unsafe_allow_html=True)


    if uploaded_file is not None:
        # Determine file extension
        file_extension = uploaded_file.name.split('.')[-1].lower()
        
        if file_extension == "xlsx":
            # Process the Excel file
            extract_excel_tables(uploaded_file)
        elif file_extension == "pdf":
            # Process the PDF file
            extract_and_combine_tables(uploaded_file)
        else:
            st.warning("Unsupported file type. Please upload an Excel or PDF file.")

if __name__ == "__main__":
    main()
