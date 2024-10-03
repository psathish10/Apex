import streamlit as st
import pandas as pd
import os
import re
import mysql.connector
from openpyxl import load_workbook
import pdfplumber
import win32com.client as win32

# Define your SQL database connection details
DB_CONFIG = {
    'host': 'localhost',  # e.g., 'localhost' or IP address
    'user': 'root',
    'password': '',
    'database': 'apex'
}

# Define the fixed column names based on your provided table structure
FIXED_COLUMNS = [
    "S.No", "BillNo/Date", "Pharmacy Name", "City", "Batch", "Expiry", "Qty",
    "Free", "Rpl", "PTS", "PTR", "MRP", "TotSales"
]

# Global variable to hold combined DataFrame for PDF
combined_df = pd.DataFrame()

# Define the column mapping for Excel
column_mapping = {
    'Customer Name': 'Stockist_Name',
    'Bill No': 'Bill_No',
    'Bill Date': 'Bill_Date',
    'Cust Code': 'Chemist_Code',
    'Product Name': 'Material_Name',
    'Qty': 'Sale_Qty',
    'Free': 'Free_Qty',
    'Amount': 'Value',
    'Batch No': 'Batch_No',
    'Address': 'Address',
    'Area Name': 'City',
    'Pin Code': 'Pin_Code',
    'Rate': 'Rate',
    'Goods Value': 'Value'
}

# Function to validate the columns against the mapping
def validate_columns(df):
    missing_columns = []
    for excel_col in column_mapping.keys():
        if excel_col not in df.columns:
            missing_columns.append(excel_col)
    return missing_columns

# Function to extract tables from Excel and display data
def extract_excel_tables(file):
    try:
        # Load the Excel file using openpyxl
        workbook = load_workbook(file)
        sheet_names = workbook.sheetnames

        # Select a sheet to extract tables from
        st.write("Available sheets in the Excel file:")
        selected_sheet = st.selectbox("Select a sheet", sheet_names)

        # Load the selected sheet
        sheet = workbook[selected_sheet]

        # Extract the tables (if any) within the selected sheet
        if sheet.tables:
            table_names = list(sheet.tables.keys())
            st.write(f"Tables in the '{selected_sheet}' sheet:")
            selected_table = st.selectbox("Select a table", table_names)

            # Extract the table range
            table_range = sheet.tables[selected_table].ref

            # Extract the cell values from the table range
            data = sheet[table_range]
            table_data = [[cell.value for cell in row] for row in data]

            # Create a DataFrame using the extracted data
            df = pd.DataFrame(table_data[1:], columns=table_data[0])  # Use the first row as column headers

            # Validate columns
            missing_columns = validate_columns(df)
            if missing_columns:
                st.warning(f"Missing columns in the uploaded table: {', '.join(missing_columns)}")
            else:
                # Display table data
                st.write(f"Data from table '{selected_table}':")
                st.dataframe(df)

                # Convert 'Bill Date' to the correct format
                df['Bill Date'] = pd.to_datetime(df['Bill Date'], format='%d/%m/%Y', errors='coerce').dt.strftime('%Y-%m-%d')

                # Insert data into MySQL
                if st.button("Insert Data into MySQL"):
                    insert_excel_data_into_db(df)
        else:
            st.write(f"No tables found in the '{selected_sheet}' sheet.")

    except Exception as e:
        st.error(f"Error occurred: {e}")

# Function to insert Excel data into the database
def insert_excel_data_into_db(df):
    try:
        # Establish database connection
        connection = mysql.connector.connect(**DB_CONFIG)
        cursor = connection.cursor()

        # Prepare SQL insert query
        insert_query = """
        INSERT INTO salesdata (Stockist_Code, Stockist_Name, Bill_No, Bill_Date, Chemist_Code, 
                                  Chemist_Name, Address, City, Pin_Code, Material_Code, 
                                  Material_Name, Batch_No, Sale_Qty, Free_Qty, Rate, Value)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """

        # Track the number of successfully inserted rows
        success_count = 0

        for index, row in df.iterrows():
            # Prepare values based on the mapping, replacing NaN with None
            values = (
                row.get('Cust Code', None),
                'KAMALAM MEDICAL CORPORATION',                                                    
                row.get('Bill No', None),
                row.get('Bill Date', None),
                'Chemist Code',  
                row.get('Customer Name', None),
                row.get('Address', None),
                row.get('Area Name', None),
                row.get('Pin Code', None),
                'Material Code',  # Replace with actual value if available
                row.get('Product Name', None),
                row.get('Batch No', None),
                row.get('Qty', None),
                row.get('Free', None),
                row.get('Rate', None),
                row.get('Amount', None)
            )
            
            # Convert NaN to None
            values = tuple(v if pd.notna(v) else None for v in values)
            
            try:
                # Insert the validated data into MySQL
                cursor.execute(insert_query, values)
                connection.commit()
                success_count += 1  # Increment the count for successful insertions
            except mysql.connector.Error as e:
                st.error(f"Error inserting row {index + 1}: {e}")

        # Show one success message after all rows are inserted
        if success_count == len(df):
            st.success(f"All {success_count} rows were inserted successfully!")
        else:
            st.warning(f"{success_count} rows were inserted successfully out of {len(df)}.")
    
    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()

def clean_and_fix_columns(data):
    df = pd.DataFrame(data)
    
    if len(df.columns) != len(FIXED_COLUMNS):
        df = df.iloc[:, :len(FIXED_COLUMNS)]
        df.columns = FIXED_COLUMNS[:len(df.columns)]
    else:
        df.columns = FIXED_COLUMNS
    
    df = df[~df.apply(lambda row: row.tolist() == FIXED_COLUMNS, axis=1)]
    df = df.dropna(how='any')
    df = df[df.apply(lambda row: all(row.astype(str).str.strip() != ''), axis=1)]
    df = df[~df.apply(lambda row: "Pharmacy Name" in row.values, axis=1)]
    
    if "BillNo/Date" in df.columns:
        df[['Bill No', 'Bill Date']] = df['BillNo/Date'].str.split('/', n=1, expand=True)
        df['Bill Date'] = df['Bill Date'].str.strip()
        df['Bill No'] = df['Bill No'].str.strip()
    
    df = df.drop(columns=['BillNo/Date'], errors='ignore')

    # Separate City and Pin Code
    if "City" in df.columns:
        df[['City', 'Pin Code']] = df['City'].str.extract(r'([A-Za-z\s]+)(\d{4,6})$')
    
    return df

# Function to associate product names by batch
def associate_product_names_by_batch(df, product_names):
    result = []
    current_product = None
    current_batch_prefix = None

    for _, row in df.iterrows():
        batch = str(row['Batch'])
        batch_prefix = batch[:3] if len(batch) >= 3 else batch

        if batch_prefix != current_batch_prefix:
            if product_names:
                current_product = product_names.pop(0)
            else:
                current_product = "Unknown Product"
            current_batch_prefix = batch_prefix

        row['Product Name'] = current_product
        result.append(row)

    return pd.DataFrame(result)

# Function to extract product name from PDF
def extract_product_name(page):
    text = page.extract_text()
    match = re.search(r"Product Name\s*:\s*(.*)", text)
    return match.group(1).strip() if match else None

# Function to extract and combine tables from PDF
def extract_and_combine_tables(pdf_file):
    global combined_df
    combined_df = pd.DataFrame()  # Reset combined DataFrame
    product_names = []

    try:
        with pdfplumber.open(pdf_file) as pdf:
            for page in pdf.pages:
                product_name = extract_product_name(page)
                if product_name:
                    product_names.append(product_name)
                
                extracted_tables = page.extract_tables()
                for table in extracted_tables:
                    df = clean_and_fix_columns(table)
                    if not df.empty:
                        combined_df = pd.concat([combined_df, df], ignore_index=True)

        if not combined_df.empty:
            combined_df = associate_product_names_by_batch(combined_df, product_names)
            st.write("Combined table from all pages with associated product names (based on Batch):")
            st.dataframe(combined_df)

            # Add a button to push data to the database
            if st.button("Insert PDF Data into Database"):
                insert_pdf_data_into_db(combined_df)
        else:
            st.write("No valid data found in the PDF tables.")
    
    except Exception as e:
        st.error(f"Error occurred while extracting tables: {str(e)}")

# Function to insert PDF data into the database
def insert_pdf_data_into_db(df):
    try:
        # Establish database connection
        connection = mysql.connector.connect(**DB_CONFIG)
        cursor = connection.cursor()

        # Define your SQL insert statement
        insert_query = """
        INSERT INTO salesdata (Stockist_Code, Stockist_Name, Bill_No, Bill_Date, Chemist_Code,
                                  Chemist_Name, Address, City, Pin_Code, Material_Code, 
                                  Material_Name, Batch_No, Sale_Qty, Free_Qty, Rate, Value)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """

        # Iterate through the DataFrame and insert each row into the database
        for index, row in df.iterrows():
            # Convert 'Bill Date' to 'yyyy-mm-dd' format
            bill_date = pd.to_datetime(row['Bill Date'], format='%d-%m-%Y', errors='coerce')
            bill_date_str = bill_date.strftime('%Y-%m-%d') if pd.notna(bill_date) else None

            data = (
                row['S.No'] if pd.notna(row.get('S.No')) else None,  # Assuming this corresponds to Stockist_Code
               "PURANI HOSPITAL SUPPLIES PVT LTD",  # Stockist Name
                row['Bill No'] if pd.notna(row.get('Bill No')) else None,
                bill_date_str,  # Use the formatted Bill Date here
                "Chemist Code",  # Replace with the actual Chemist Code
                row['Pharmacy Name'] if pd.notna(row.get('Pharmacy Name')) else None,  # Stockist Name
                row['Address'] if pd.notna(row.get('Address')) else None,
                row['City'] if pd.notna(row.get('City')) else None,
                row['Pin Code'] if pd.notna(row.get('Pin Code')) else None,
                "Material Code",  # Replace with the actual Material Code
                row['Product Name'] if pd.notna(row.get('Product Name')) else None,
                row['Batch'] if pd.notna(row.get('Batch')) else None,
                row['Qty'] if pd.notna(row.get('Qty')) else None,  # Assuming Sale_Qty corresponds to Qty
                row['Free'] if pd.notna(row.get('Free')) else None,  # Assuming Free_Qty corresponds to Free
                row['Rate'] if pd.notna(row.get('Rate')) else None,  # Rate mapped to PTR
                row['TotSales'] if pd.notna(row.get('TotSales')) else None  # Value mapped to TotSales
            )
            cursor.execute(insert_query, data)

        # Commit the transaction
        connection.commit()
        st.success("Data has been successfully inserted into the database.")

    except mysql.connector.Error as err:
        st.error(f"Error: {err}")
    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()

# Function to extract text from .doc file using pywin32
def extract_text_from_doc(file):
    try:
        # Save the uploaded file to a temporary path
        temp_doc_path = os.path.join(os.getcwd(), "temp.doc")
        with open(temp_doc_path, "wb") as f:
            f.write(file.read())

        # Initialize Word application
        word = win32.Dispatch("Word.Application")
        word.Visible = False

        # Open the .doc file using the absolute path
        doc = word.Documents.Open(temp_doc_path)

        # Extract the text from the document
        full_text = doc.Content.Text

        # Close the document
        doc.Close(False)
        word.Quit()

        # Delete the temporary file
        os.remove(temp_doc_path)

        return full_text

    except Exception as e:
        st.error(f"Error occurred while extracting text: {e}")
        return None

# Function to process and extract structured data from the Word document
def process_structured_data_with_product_names(text):
    lines = text.splitlines()

    # Regular expression to match rows with customer name, pin, qty, and amounts
    row_pattern = re.compile(r"([A-Za-z\s.]+)\s+(\d{6})\s+(-?\d+)\s+(-?\d+\.\d+)\s+(-?\d+\.\d+)")

    # Define the stockist name
    stockist_name = "LIFECARE PHARMA PRIVATE LIMITED"

    # List to hold the extracted data
    data = []
    current_product_name = None
    next_line_is_product = False

    for line in lines:
        line = line.strip()

        # Check for the hyphen separator line
        if line.startswith('---'):
            next_line_is_product = True  # The next line is the product name
            continue

        # If the next line is product name, store it and reset the flag
        if next_line_is_product:
            current_product_name = line
            next_line_is_product = False
            continue

        # Check if the line matches the row pattern
        match = row_pattern.match(line)

        if match:
            # Extract row details and append the current product name and stockist name
            customer_name, pin, qty, gross_amt, net_amt = match.groups()
            data.append({
                "Stockist Name": stockist_name,  # Add stockist name to each row
                "Product Name": current_product_name,
                "Customer Name": customer_name.strip(),
                "Pin": pin.strip(),
                "Qty": int(qty),
                "Gross Amt": float(gross_amt),
                "Net Amt": float(net_amt)
            })

    # Convert the list to a DataFrame
    df = pd.DataFrame(data)
    
    return df

# Function to insert extracted data from Word document into MySQL database
def insert_data_from_word_into_db(df):
    try:
        # Establish a database connection
        connection = mysql.connector.connect(**DB_CONFIG)
        cursor = connection.cursor()

        # Prepare SQL insert query
        insert_query = """
        INSERT INTO salesdata (Stockist_Code, Stockist_Name, Bill_No, Bill_Date, Chemist_Code, 
                                  Chemist_Name, Address, City, Pin_Code, Material_Code, 
                                  Material_Name, Batch_No, Sale_Qty, Free_Qty, Rate, Value)
        VALUES (%s, %s, %s, CURDATE(), %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """

        for index, row in df.iterrows():
            values = (
                'CODE',  # Stockist_Code
                row['Stockist Name'],  # Stockist_Name
                'BILL_NO',  # Bill_No
                'CHEMIST_CODE',  # Chemist_Code
                row['Customer Name'],  # Chemist_Name
                'ADDRESS',  # Address
                'CITY',  # City
                row['Pin'],  # Pin_Code
                'MATERIAL_CODE',  # Material_Code
                row['Product Name'],  # Material_Name
                'BATCH_NO',  # Batch_No
                row['Qty'],  # Sale_Qty
                0,  # Free_Qty
                row['Gross Amt'],  # Rate (Gross Amt)
                row['Net Amt'],  # Value (Net Amt)
            )

            cursor.execute(insert_query, values)

        # Commit the transaction
        connection.commit()
        st.success("All data has been successfully inserted into the database.")
        
    except mysql.connector.Error as err:
        st.error(f"Error inserting data into MySQL: {err}")
    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()

def main():
    st.image("toolfe.webp", width=200)
    st.title("Data Extraction and Insertion Application")
    
    st.sidebar.title("Upload Options")
    uploaded_file = st.sidebar.file_uploader("Upload a file (Excel, PDF, Word)", type=["xlsx", "xls", "pdf", "doc"])

    if uploaded_file:
        file_extension = uploaded_file.name.split('.')[-1].lower()
        
        with st.spinner("Processing your file..."):
            if file_extension in ["xlsx", "xls"]:
                st.write("Processing Excel file...")
                df = extract_excel_tables(uploaded_file)  # Modify this function to return the DataFrame

                if df is not None:  # Check if the DataFrame is not empty
                    # Place the insert button above the dataframe
                    if st.button("Insert Data into Database"):
                        insert_excel_data_into_db(df)
                    
                    # Display the dataframe
                    st.dataframe(df)

            elif file_extension == "pdf":
                st.write("Processing PDF file...")
                df = extract_and_combine_tables(uploaded_file)  # Modify this function to return the DataFrame

                if df is not None:  # Check if the DataFrame is not empty
                    # Place the insert button above the dataframe
                    if st.button("Insert PDF Data into Database"):
                        insert_pdf_data_into_db(df)
                    
                    # Display the dataframe
                    st.dataframe(df)

            elif file_extension == "doc":
                st.write("Processing Word document...")
                extracted_text = extract_text_from_doc(uploaded_file)
                if extracted_text:
                    df = process_structured_data_with_product_names(extracted_text)
                    st.write("Extracted Table from Word Document:")
                    
                    # Place the insert button above the dataframe
                    if st.button("Insert Data from Word Document into Database"):
                        insert_data_from_word_into_db(df)
                    
                    # Display the dataframe
                    st.dataframe(df)

if __name__ == "__main__":
    main()
