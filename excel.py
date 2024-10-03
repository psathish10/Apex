# excel.py

import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import mysql.connector
from db_config import DB_CONFIG

# Define the column mapping for Excel
column_mapping = {
    'Customer Name': 'Stockist_Name',
    'Bill No': 'Bill_No',
    'Bill Date': 'Bill_Date',
    'Cust Code': 'Chemist_Code',
    'Product Name': 'Material_Name',
    'Qty': 'Sale_Qty',
    'Free': 'Free_Qty',
    'Amount': 'Value',
    'Batch No': 'Batch_No',
    'Address': 'Address',
    'Area Name': 'City',
    'Pin Code': 'Pin_Code',
    'Rate': 'Rate',
    'Goods Value': 'Value'
}

# Function to validate the columns against the mapping
def validate_columns(df):
    missing_columns = [col for col in column_mapping.keys() if col not in df.columns]
    return missing_columns

# Extract tables from Excel and insert data
def extract_excel_tables(file):
    try:
        workbook = load_workbook(file)
        sheet_names = workbook.sheetnames

        st.write("Available sheets in the Excel file:")
        selected_sheet = st.selectbox("Select a sheet", sheet_names)
        sheet = workbook[selected_sheet]

        if sheet.tables:
            table_names = list(sheet.tables.keys())
            selected_table = st.selectbox("Select a table", table_names)
            table_range = sheet.tables[selected_table].ref
            data = sheet[table_range]
            table_data = [[cell.value for cell in row] for row in data]
            df = pd.DataFrame(table_data[1:], columns=table_data[0])

            missing_columns = validate_columns(df)
            if missing_columns:
                st.warning(f"Missing columns: {', '.join(missing_columns)}")
            else:
                df['Bill Date'] = pd.to_datetime(df['Bill Date'], format='%d/%m/%Y', errors='coerce').dt.strftime('%Y-%m-%d')
                if st.button("Insert Excel Data"):
                    insert_excel_data_into_db(df)
        else:
            st.write("No tables found in the sheet.")
    except Exception as e:
        st.error(f"Error: {e}")

# Insert data into MySQL from Excel
def insert_excel_data_into_db(df):
    try:
        connection = mysql.connector.connect(**DB_CONFIG)
        cursor = connection.cursor()

        insert_query = """
        INSERT INTO salesdata (Stockist_Code, Stockist_Name, Bill_No, Bill_Date, Chemist_Code, Chemist_Name, Address, City, Pin_Code, Material_Code, Material_Name, Batch_No, Sale_Qty, Free_Qty, Rate, Value)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """

        success_count = 0
        for _, row in df.iterrows():
            values = (
                row.get('Cust Code', None),
                'KAMALAM MEDICAL CORPORATION',
                row.get('Bill No', None),
                row.get('Bill Date', None),
                'Chemist Code',
                row.get('Customer Name', None),
                row.get('Address', None),
                row.get('Area Name', None),
                row.get('Pin Code', None),
                'Material Code',
                row.get('Product Name', None),
                row.get('Batch No', None),
                row.get('Qty', None),
                row.get('Free', None),
                row.get('Rate', None),
                row.get('Amount', None)
            )
            values = tuple(v if pd.notna(v) else None for v in values)
            try:
                cursor.execute(insert_query, values)
                connection.commit()
                success_count += 1
            except mysql.connector.Error as e:
                st.error(f"Error inserting row: {e}")

        st.success(f"Inserted {success_count} rows successfully.")
    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()
